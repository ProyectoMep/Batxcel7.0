"""Formato visual del Excel de salida: encabezados rojos, bordes,
pestañas de resumen coloreadas y orden de hojas por categoría.
Reemplaza core/formato.py del sistema original.
"""
from openpyxl.styles import Font, PatternFill, Border, Side


def aplicar_formato_encabezados(workbook, sheet_name: str = None) -> None:
    """Encabezados con fondo rojo y texto blanco; bordes en celdas con datos."""
    font_blanco = Font(color="FFFFFF", bold=True)
    fill_rojo = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    lado = Side(style="thin", color="000000")
    borde = Border(left=lado, right=lado, top=lado, bottom=lado)

    hojas = [workbook[sheet_name]] if sheet_name else workbook.worksheets
    for ws in hojas:
        max_row, max_col = ws.max_row, ws.max_column

        for r in range(ws.max_row, 0, -1):
            if any(ws.cell(row=r, column=c).value not in (None, "")
                   for c in range(1, ws.max_column + 1)):
                max_row = r
                break
        for c in range(ws.max_column, 0, -1):
            if any(ws.cell(row=r, column=c).value not in (None, "")
                   for r in range(1, max_row + 1)):
                max_col = c
                break

        if max_row == 0 or max_col == 0:
            continue

        for c in range(1, max_col + 1):
            celda = ws.cell(row=1, column=c)
            celda.font = font_blanco
            celda.fill = fill_rojo
            celda.border = borde

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).border = borde


def colorear_pestanas_resumen(workbook) -> None:
    for ws in workbook.worksheets:
        if ws.title.startswith("Resumen"):
            ws.sheet_properties.tabColor = "FF6B6B"


def reordenar_hojas_por_categoria(workbook, categorias_por_hoja: dict) -> None:
    """Coloca el resumen de cada categoría al final de sus hojas."""
    hojas_por_categoria, hojas_resumen, sin_categoria = {}, [], []

    for ws in workbook.worksheets:
        if ws.title.startswith("Resumen"):
            hojas_resumen.append(ws)
        elif ws.title in categorias_por_hoja:
            hojas_por_categoria.setdefault(categorias_por_hoja[ws.title], []).append(ws)
        else:
            sin_categoria.append(ws)

    nuevo_orden = []
    for categoria, hojas in hojas_por_categoria.items():
        nuevo_orden.extend(hojas)
        nombre_resumen = f"Resumen_{categoria.replace(' ', '_')}"
        for r in hojas_resumen:
            if r.title == nombre_resumen:
                nuevo_orden.append(r)
                break

    nuevo_orden.extend(sin_categoria)
    for r in hojas_resumen:
        if r not in nuevo_orden:
            nuevo_orden.append(r)

    for idx, ws in enumerate(nuevo_orden):
        workbook.move_sheet(ws, offset=idx - workbook.index(ws))