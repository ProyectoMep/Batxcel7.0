"""Usuarios PROSPERA sin MFA: cruza la hoja MFA del Resumen Cyber BSNC
de una fecha con el archivo Cloud_Authentication que subes, y softerra,
dejando solo usuarios 'Customer Sales & Support' sin ningún método de
autenticación. Puerto exacto de prospera.py del sistema original —
Prospera siempre trabaja sobre BSNC, igual que en el original.
"""
import re
from datetime import date
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

from config.settings import OUTPUT_DIR, TEMPLATES_MAESTROS_DIR, INPUT_DIR, PROCESADOS_DIR

MESES = ["", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO",
         "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]

COLUMNAS_VACIAS = ["registered methods", "mfa methods", "passwordless methods",
                   "non-phishing resistant methods"]

COLUMNAS_PLANTILLA = ["Company", "User Principal Name", "User Email Address",
                      "Nombre", "Entidad", "Registered Methods", "MFA Methods",
                      "Passwordless Methods", "Non-Phishing Resistant Methods",
                      "Block Non-Phishing Resistant Methods", "Account Enabled"]


def _es_vacio(serie: pd.Series) -> pd.Series:
    s = serie.astype(str).str.strip().str.lower()
    return serie.isna() | s.isin(["", "nan", "none", "n/a", "na", "[]", "[ ]"])


def _ruta_plantilla() -> Path | None:
    if not TEMPLATES_MAESTROS_DIR.exists():
        return None
    candidatos = list(TEMPLATES_MAESTROS_DIR.glob("PROSPERA*plantilla*.xlsx")) + \
                 list(TEMPLATES_MAESTROS_DIR.glob("PROSPERA*PLANTILLA*.xlsx"))
    return candidatos[0] if candidatos else None


def _buscar_softerra(fecha_resumen: str) -> Path | None:
    """Busca softerra.xlsx: output/<fecha>/, luego el más reciente de
    output/, luego procesados/, luego input/."""
    directo = OUTPUT_DIR / fecha_resumen / "softerra.xlsx"
    if directo.exists():
        return directo
    if OUTPUT_DIR.exists():
        for carpeta in sorted(OUTPUT_DIR.iterdir(), reverse=True):
            f = carpeta / "softerra.xlsx"
            if f.exists():
                return f
    if PROCESADOS_DIR.exists():
        for carpeta in sorted(PROCESADOS_DIR.iterdir(), reverse=True):
            f = carpeta / "softerra.xlsx"
            if f.exists():
                return f
    f = INPUT_DIR / "softerra.xlsx"
    return f if f.exists() else None


def generar_prospera(fecha_resumen: str, ruta_cloud, nombre: str = None) -> dict:
    """
    1) Toma los IDs de la hoja MFA del Resumen Cyber BSNC de fecha_resumen.
    2) Del archivo Cloud cargado, filtra usuarios con los 4 métodos vacíos.
    3) Mantiene solo los que están en ambos (resumen ∩ cloud).
    4) Cruza con softerra y conserva solo títulos 'Customer Sales & Support'
       (Nombre = description, Entidad = title).
    5) Llena la plantilla PROSPERA y la guarda en output/<fecha_resumen>/prospera/.
    """
    ruta_resumen = OUTPUT_DIR / fecha_resumen / "Resumen Cyber BSNC.xlsx"
    if not ruta_resumen.exists():
        return {"ok": False, "error": f"No existe Resumen Cyber BSNC en {fecha_resumen}"}

    ruta_plantilla = _ruta_plantilla()
    if ruta_plantilla is None:
        return {"ok": False,
                "error": "No se encontró la plantilla PROSPERA en templates_maestros/maestros/"}

    ruta_softerra = _buscar_softerra(fecha_resumen)
    if ruta_softerra is None:
        return {"ok": False, "error": "No se encontró softerra.xlsx en el histórico"}

    # 1) IDs de la hoja MFA del resumen (ya vienen sin @dominio)
    xl = pd.ExcelFile(ruta_resumen)
    hoja_mfa = next((h for h in xl.sheet_names if h.strip().lower() == "mfa"), None)
    if hoja_mfa is None:
        return {"ok": False, "error": "El Resumen seleccionado no tiene hoja MFA"}
    df_mfa = xl.parse(hoja_mfa)
    df_mfa.columns = [str(c).strip().lower() for c in df_mfa.columns]
    col_id = next((c for c in df_mfa.columns if "user principal name" in c), None)
    if col_id is None:
        return {"ok": False, "error": "La hoja MFA no tiene columna User Principal Name"}
    ids_resumen = set(df_mfa[col_id].dropna().astype(str)
                      .str.split("@").str[0].str.strip().str.lower())

    # 2) Cloud: usuarios con los 4 métodos vacíos
    cloud = pd.read_excel(ruta_cloud, dtype=str)
    cols_norm = {str(c).strip().lower(): c for c in cloud.columns}
    faltan = [c for c in COLUMNAS_VACIAS + ["user principal name"] if c not in cols_norm]
    if faltan:
        return {"ok": False, "error": f"El archivo Cloud no tiene columnas: {faltan}"}
    mask = pd.Series(True, index=cloud.index)
    for c in COLUMNAS_VACIAS:
        mask &= _es_vacio(cloud[cols_norm[c]])
    sin_metodos = cloud[mask].copy()

    # 3) Intersección con el resumen
    sin_metodos["_usuario"] = (sin_metodos[cols_norm["user principal name"]]
                               .astype(str).str.split("@").str[0]
                               .str.strip().str.lower())
    coincide = sin_metodos[sin_metodos["_usuario"].isin(ids_resumen)].copy()

    # 4) Softerra: solo Customer Sales & Support
    soft = pd.read_excel(ruta_softerra, dtype=str)
    soft.columns = [str(c).strip().lower() for c in soft.columns]
    soft["_key"] = soft["name"].astype(str).str.strip().str.lower()
    soft = soft.drop_duplicates("_key")
    final = coincide.merge(soft[["_key", "description", "title"]],
                           left_on="_usuario", right_on="_key", how="left")
    final = final[final["title"].astype(str)
                  .str.contains("Customer Sales & Support", case=False, na=False)]

    # 5) Llenar la plantilla conservando su formato
    wb = load_workbook(ruta_plantilla)
    ws = wb.active
    encabezados = {str(ws.cell(row=1, column=c).value).strip(): c
                  for c in range(1, ws.max_column + 1)}

    def valor(fila, col_plantilla):
        if col_plantilla == "Nombre":
            v = fila.get("description")
        elif col_plantilla == "Entidad":
            v = fila.get("title")
        else:
            v = fila.get(cols_norm.get(col_plantilla.lower(), ""), None)
            if v is None:
                v = fila.get(col_plantilla)
        return "" if pd.isna(v) else str(v)

    for i, (_, fila) in enumerate(final.iterrows(), start=2):
        for col_plantilla, col_idx in encabezados.items():
            if col_plantilla in COLUMNAS_PLANTILLA:
                ws.cell(row=i, column=col_idx, value=valor(fila, col_plantilla))

    # Nombre del archivo (por defecto con la fecha del resumen en español)
    if not nombre:
        d = date.fromisoformat(fecha_resumen)
        nombre = f"PROSPERA_SIN_MFA_DESDE_EL_{d.day}_DE_{MESES[d.month]}_{d.year}"
    nombre = re.sub(r'[\\/:*?"<>|]+', "_", nombre.strip())

    carpeta = OUTPUT_DIR / fecha_resumen / "prospera"
    carpeta.mkdir(parents=True, exist_ok=True)
    ruta_salida = carpeta / f"{nombre}.xlsx"
    wb.save(ruta_salida)
    wb.close()

    return {"ok": True, "archivo": ruta_salida.name, "fecha": fecha_resumen,
            "sin_metodos": int(len(sin_metodos)),
            "en_resumen": int(len(coincide)), "finales": int(len(final))}


def listar_prospera() -> list:
    resultado = []
    if not OUTPUT_DIR.exists():
        return resultado
    for carpeta in sorted(OUTPUT_DIR.iterdir(), reverse=True):
        sub = carpeta / "prospera"
        if sub.is_dir():
            for f in sorted(sub.glob("*.xlsx")):
                resultado.append({"fecha": carpeta.name, "nombre": f.stem,
                                  "xlsx": f"prospera/{f.name}"})
    return resultado